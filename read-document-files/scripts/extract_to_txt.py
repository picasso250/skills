#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import shutil
import sys
from pathlib import Path
from typing import Iterable
from zipfile import ZipFile
import xml.etree.ElementTree as ET


W_NS = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
X_NS = {
    "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}
P_NS = {
    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
    "p": "http://schemas.openxmlformats.org/presentationml/2006/main",
}


def ensure_utf8_stdio() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")


def clean_text(value: str) -> str:
    value = value.replace("\r\n", "\n").replace("\r", "\n")
    value = re.sub(r"[ \t]+\n", "\n", value)
    value = re.sub(r"\n{3,}", "\n\n", value)
    return value.strip()


def read_xml_from_zip(doc: ZipFile, member: str) -> ET.Element | None:
    try:
        with doc.open(member) as handle:
            return ET.parse(handle).getroot()
    except KeyError:
        return None


def extract_docx_paragraph(paragraph: ET.Element) -> str:
    texts = [node.text for node in paragraph.findall(".//w:t", W_NS) if node.text]
    return "".join(texts).strip()


def extract_docx_table(table: ET.Element) -> list[str]:
    rows: list[str] = []
    for row in table.findall("./w:tr", W_NS):
        cells: list[str] = []
        for cell in row.findall("./w:tc", W_NS):
            cell_parts = [
                extract_docx_paragraph(paragraph)
                for paragraph in cell.findall("./w:p", W_NS)
            ]
            cells.append(" / ".join(part for part in cell_parts if part))
        if any(cells):
            rows.append(" | ".join(cells))
    return rows


def extract_docx(path: Path, assets_dir: Path | None, metadata: dict) -> str:
    parts: list[str] = []
    images: list[str] = []

    with ZipFile(path) as doc:
        root = read_xml_from_zip(doc, "word/document.xml")
        if root is None:
            raise ValueError("word/document.xml not found")
        body = root.find("./w:body", W_NS)
        if body is None:
            raise ValueError("w:body not found")

        for child in list(body):
            if child.tag == f"{{{W_NS['w']}}}p":
                paragraph_text = extract_docx_paragraph(child)
                if paragraph_text:
                    parts.append(paragraph_text)
            elif child.tag == f"{{{W_NS['w']}}}tbl":
                rows = extract_docx_table(child)
                if rows:
                    parts.append("[Table]")
                    parts.extend(rows)

        media_members = [
            name for name in doc.namelist() if name.startswith("word/media/") and not name.endswith("/")
        ]
        if media_members and assets_dir is not None:
            image_dir = assets_dir / "images"
            image_dir.mkdir(parents=True, exist_ok=True)
            for member in media_members:
                target = image_dir / Path(member).name
                with doc.open(member) as src, target.open("wb") as dst:
                    shutil.copyfileobj(src, dst)
                images.append(str(target))

    if images:
        metadata["images"] = images
        metadata["guidance"] = (
            "This DOCX contained embedded images. Review the exported image files with a vision-capable step "
            "if charts, screenshots, or scanned inserts may affect the answer."
        )

    return clean_text("\n\n".join(parts))


def column_name(index: int) -> str:
    name = []
    while index > 0:
        index, rem = divmod(index - 1, 26)
        name.append(chr(65 + rem))
    return "".join(reversed(name))


def extract_xlsx_with_openpyxl(path: Path) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(path, data_only=True, read_only=True)
    sections: list[str] = []
    for sheet in workbook.worksheets:
        sections.append(f"[Sheet] {sheet.title}")
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            sections.append("(empty)")
            sections.append("")
            continue
        header = [str(value).strip() if value is not None else "" for value in rows[0]]
        use_header = any(header)
        for row in rows[1:] if use_header else rows:
            values = ["" if value is None else str(value).strip() for value in row]
            if not any(values):
                continue
            if use_header:
                pairs = [
                    f"{header[idx] or column_name(idx + 1)}={value}"
                    for idx, value in enumerate(values)
                    if value
                ]
                sections.append(" ; ".join(pairs))
            else:
                sections.append("\t".join(values).rstrip())
        sections.append("")
    return clean_text("\n".join(sections))


def extract_xlsx_raw(path: Path) -> str:
    with ZipFile(path) as doc:
        shared_strings_root = read_xml_from_zip(doc, "xl/sharedStrings.xml")
        shared_strings: list[str] = []
        if shared_strings_root is not None:
            for item in shared_strings_root.findall("./main:si", X_NS):
                text = "".join(node.text or "" for node in item.findall(".//main:t", X_NS))
                shared_strings.append(text)

        workbook_root = read_xml_from_zip(doc, "xl/workbook.xml")
        if workbook_root is None:
            raise ValueError("xl/workbook.xml not found")

        rels_root = read_xml_from_zip(doc, "xl/_rels/workbook.xml.rels")
        rel_map: dict[str, str] = {}
        if rels_root is not None:
            for rel in rels_root:
                rel_id = rel.attrib.get("Id")
                target = rel.attrib.get("Target")
                if rel_id and target:
                    rel_map[rel_id] = "xl/" + target.lstrip("/")

        sections: list[str] = []
        for sheet in workbook_root.findall("./main:sheets/main:sheet", X_NS):
            name = sheet.attrib.get("name", "Sheet")
            rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
            if not rel_id or rel_id not in rel_map:
                continue
            sheet_root = read_xml_from_zip(doc, rel_map[rel_id])
            sections.append(f"[Sheet] {name}")
            if sheet_root is None:
                sections.append("(unreadable)")
                sections.append("")
                continue
            row_lines: list[list[str]] = []
            for row in sheet_root.findall(".//main:sheetData/main:row", X_NS):
                values: list[str] = []
                for cell in row.findall("./main:c", X_NS):
                    cell_type = cell.attrib.get("t")
                    value_node = cell.find("./main:v", X_NS)
                    value = value_node.text if value_node is not None and value_node.text else ""
                    if cell_type == "s" and value:
                        idx = int(value)
                        value = shared_strings[idx] if idx < len(shared_strings) else value
                    values.append(value.strip())
                if any(values):
                    row_lines.append(values)
            for values in row_lines:
                sections.append("\t".join(values))
            sections.append("")
    return clean_text("\n".join(sections))


def extract_xlsx(path: Path, metadata: dict) -> str:
    try:
        return extract_xlsx_with_openpyxl(path)
    except Exception as exc:
        metadata.setdefault("warnings", []).append(
            f"openpyxl path failed, fell back to raw OOXML parsing: {exc}"
        )
        return extract_xlsx_raw(path)


def extract_pptx(path: Path) -> str:
    with ZipFile(path) as doc:
        slide_names = sorted(
            name for name in doc.namelist() if name.startswith("ppt/slides/slide") and name.endswith(".xml")
        )
        sections: list[str] = []
        for index, slide_name in enumerate(slide_names, start=1):
            root = read_xml_from_zip(doc, slide_name)
            if root is None:
                continue
            sections.append(f"[Slide {index}]")
            slide_text = [
                node.text.strip()
                for node in root.findall(".//a:t", P_NS)
                if node.text and node.text.strip()
            ]
            sections.extend(slide_text or ["(no text)"])
            sections.append("")
    return clean_text("\n".join(sections))


def extract_pdf(path: Path, metadata: dict) -> str:
    reader = None
    import_error = None
    for module_name in ("pypdf", "PyPDF2"):
        try:
            module = __import__(module_name)
            reader = module.PdfReader(str(path))
            metadata["pdf_library"] = module_name
            break
        except Exception as exc:
            import_error = exc
    if reader is None:
        raise RuntimeError(
            "PDF extraction requires pypdf or PyPDF2. Install one of them in this Python environment."
        ) from import_error

    pages: list[str] = []
    for idx, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception as exc:
            metadata.setdefault("warnings", []).append(f"Failed to read PDF page {idx}: {exc}")
            text = ""
        pages.append(f"[Page {idx}]\n{text.strip()}".strip())
    return clean_text("\n\n".join(pages))


def write_outputs(
    source_path: Path,
    output_dir: Path,
    text: str,
    metadata: dict,
    echo_stdout: bool,
) -> None:
    output_dir.mkdir(parents=True, exist_ok=True)
    txt_path = output_dir / f"{source_path.stem}.txt"
    meta_path = output_dir / f"{source_path.stem}.meta.json"
    txt_path.write_text(text + ("\n" if text else ""), encoding="utf-8")
    meta_path.write_text(json.dumps(metadata, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
    print(f"Wrote text: {txt_path}")
    print(f"Wrote metadata: {meta_path}")
    if metadata.get("images"):
        print(f"Extracted {len(metadata['images'])} image(s) from DOCX.")
    if echo_stdout:
        print("\n--- BEGIN TEXT ---\n")
        print(text)
        print("\n--- END TEXT ---")


def extract_file(path: Path, output_dir: Path, echo_stdout: bool) -> int:
    suffix = path.suffix.lower()
    metadata: dict = {
        "source": str(path),
        "format": suffix.lstrip("."),
        "warnings": [],
    }
    assets_dir = output_dir / f"{path.stem}_assets"

    try:
        if suffix == ".docx":
            text = extract_docx(path, assets_dir, metadata)
        elif suffix == ".xlsx":
            text = extract_xlsx(path, metadata)
        elif suffix == ".pptx":
            text = extract_pptx(path)
        elif suffix == ".pdf":
            text = extract_pdf(path, metadata)
        else:
            raise ValueError(
                f"Unsupported extension: {suffix or '(none)'}. Supported: .pdf .docx .xlsx .pptx"
            )
    except Exception as exc:
        metadata.setdefault("warnings", []).append(str(exc))
        text = ""
        write_outputs(path, output_dir, text, metadata, echo_stdout=False)
        print(f"Extraction failed: {exc}", file=sys.stderr)
        return 1

    write_outputs(path, output_dir, text, metadata, echo_stdout)
    return 0


def parse_args(argv: Iterable[str]) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract readable text from document files into UTF-8 .txt output."
    )
    parser.add_argument("path", help="Path to the source file")
    parser.add_argument(
        "--output-dir",
        help="Directory for generated .txt/.meta.json files. Defaults to <source-dir>\\_extracted",
    )
    parser.add_argument(
        "--stdout",
        action="store_true",
        help="Print extracted text to stdout after writing files",
    )
    return parser.parse_args(list(argv))


def main(argv: Iterable[str] | None = None) -> int:
    ensure_utf8_stdio()
    args = parse_args(argv or sys.argv[1:])
    source_path = Path(args.path).expanduser().resolve()
    if not source_path.exists():
        print(f"File not found: {source_path}", file=sys.stderr)
        return 1
    output_dir = (
        Path(args.output_dir).expanduser().resolve()
        if args.output_dir
        else source_path.parent / "_extracted"
    )
    return extract_file(source_path, output_dir, args.stdout)


if __name__ == "__main__":
    raise SystemExit(main())
